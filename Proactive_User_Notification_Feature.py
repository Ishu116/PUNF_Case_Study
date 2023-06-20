import getpass
import smtplib
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


class LinkedInScraper:
    def __init__(self, chromedriver_path):
        self.driver = self._create_driver(chromedriver_path)

    def _create_driver(self, chromedriver_path):
        service = Service(chromedriver_path)
        return webdriver.Chrome(service=service)

    def _login(self, username, password):
        self.driver.get('https://www.linkedin.com')
        # Find the email input field and enter the username
        email_input = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.ID, 'session_key')))
        email_input.send_keys(username)
        # Find the password input field and enter the password
        password_input = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.ID, 'session_password')))
        password_input.send_keys(password)
        # Click the sign-in button
        sign_in_button = self.driver.find_element(By.XPATH, '//button[@type="submit"]')
        sign_in_button.click()

    def _get_unread_counts(self):
        # Find the element for unread messages
        unread_messages_element = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, '[href="https://www.linkedin.com/messaging/?"]')))
        unread_messages_count = 0
        try:
            # Extract the count of unread messages
            unread_messages = unread_messages_element.find_element(By.CLASS_NAME, 'ember-view')
            unread_messages_show_element = unread_messages.find_element(By.CLASS_NAME, 'notification-badge--show')
            unread_messages_count_element = unread_messages_show_element.find_element(By.CLASS_NAME, 'notification-badge__count')
            unread_messages_count = int(unread_messages_count_element.text)
        except Exception as e:
            pass

        # Find the element for notifications
        notifications_element = WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, '[href="https://www.linkedin.com/notifications/?"]')))
        notifications_count = 0
        try:
            # Extract the count of notifications
            notifications_badge = notifications_element.find_element(By.CLASS_NAME, 'ember-view')
            notifications_show = notifications_badge.find_element(By.CLASS_NAME, 'notification-badge--show')
            notifications = notifications_show.find_element(By.CLASS_NAME, 'notification-badge__count')
            notifications_count = int(notifications.text)
        except Exception as e:
            pass

        return unread_messages_count, notifications_count

    def scrape_unread_data(self, username, password):
        try:
            self._login(username, password)
            unread_messages_count, notifications_count = self._get_unread_counts()
            return unread_messages_count, notifications_count
        except Exception as e:
            print("An error occurred:", e)
        finally:
            self.driver.quit()


# def generate_email_body(current_data, previous_data):
#     body = ''
#     if previous_data:
#         # Prepare the email body with current and previous data
#         body = {
#             "Previous Unread Messages:": previous_data["unread_messages"],
#             "Previous Unread Notifications:": previous_data["unread_notifications"],
#             "Current Unread Messages:": current_data[0],
#             "Current Unread Notifications:": current_data[1],
#             "Comparison:": findComparison(current_data, previous_data)
#         }
#     else:
#         # Prepare the email body with only current data
#         body = {
#             "Current Unread Messages:": current_data[0],
#             "Current Unread Notifications:": current_data[1]
#         }
#     return body

def generate_email_body(current_data, previous_data):
    css_styles = '''
    <style>
        body {
            font-family: Arial, sans-serif;
            padding: 20px;
            background-color: #f5f5f5;
        }

        h2 {
            color: #0073b1;
            margin-bottom: 20px;
        }

        h3 {
            color: #333333;
            margin-bottom: 10px;
        }

        ul {
            margin-top: 0;
            margin-bottom: 20px;
            padding-left: 20px;
        }

        li {
            margin-bottom: 5px;
        }

        .highlight {
            color: #0073b1;
            font-weight: bold;
        }
    </style>
    '''

    if previous_data:
        # Prepare the email body with current and previous data
        body = f'''
        <html>
        <head>{css_styles}</head>
        <body>
            <h2>LinkedIn Unread Messages and Notifications</h2>
            <h3>Previous Unread Messages: {previous_data["unread_messages"]}</h3>
            <h3>Previous Unread Notifications: {previous_data["unread_notifications"]}</h3>
            <h3>Current Unread Messages: {current_data[0]}</h3>
            <h3>Current Unread Notifications: {current_data[1]}</h3>
            <h3>Comparison:</h3>
            <ul>
                <li><span class="highlight">Unread Messages:</span> {current_data[0] - previous_data["unread_messages"]}</li>
                <li><span class="highlight">Unread Notifications:</span> {current_data[1] - previous_data["unread_notifications"]}</li>
            </ul>
        </body>
        </html>
        '''
    else:
        # Prepare the email body with only current data
        body = f'''
        <html>
        <head>{css_styles}</head>
        <body>
            <h2>LinkedIn Unread Messages and Notifications</h2>
            <h3>Current Unread Messages: {current_data[0]}</h3>
            <h3>Current Unread Notifications: {current_data[1]}</h3>
        </body>
        </html>
        '''
    return body



# def send_email(sender_email, sender_password, recipient_email, SMTP_SERVER, SMTP_PORT, subject, body):
#     message = f'Subject: {subject}\n\n{body}'

#     with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
#         server.starttls()
#         server.login(sender_email, sender_password)
#         server.sendmail(sender_email, recipient_email, message.encode('utf-8'))

def send_email(sender_email, sender_password, recipient_email, SMTP_SERVER, SMTP_PORT, subject, body):
    # Create a multipart message object
    message = MIMEMultipart("alternative")
    message["Subject"] = subject
    message["From"] = sender_email
    message["To"] = recipient_email

    # Create HTML content for the email
    html_content = body

    # Attach the HTML content to the message
    message.attach(MIMEText(html_content, "html"))

    # Setup the SMTP server and send the email
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, recipient_email, message.as_string())


def update_excel_data(filename, data):
    workbook = load_workbook(filename)
    sheet = workbook.active
    sheet.append(data)
    workbook.save(filename)


def retrieve_previous_data(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active
    return {
        "unread_messages": sheet["C"][-1].value,
        "unread_notifications": sheet["D"][-1].value
    }


def findComparison(current_data, previous_data):
    if previous_data['unread_messages'] < current_data[0] and previous_data['unread_notifications'] < current_data[1]:
        return {
            "Unread Messages:": current_data[0] - previous_data["unread_messages"],
            "Unread Notifications:": current_data[1] - previous_data["unread_notifications"]
        }
    elif previous_data['unread_messages'] < current_data[0] and previous_data['unread_notifications'] > current_data[1]:
        return {
            "Unread Messages:": current_data[0] - previous_data["unread_messages"],
            "Unread Notifications:": current_data[1] - previous_data["unread_notifications"]
        }
    elif previous_data['unread_messages'] > current_data[0] and previous_data['unread_notifications'] < current_data[1]:
        return {
            "Unread Messages:": current_data[0] - previous_data["unread_messages"],
            "Unread Notifications:": current_data[1] - previous_data["unread_notifications"]
        }
    else:
        return {
            "Unread Messages:": current_data[0] - previous_data["unread_messages"],
            "Unread Notifications:": current_data[1] - previous_data["unread_notifications"]
        }


if __name__ == '__main__':
    # Specify the path to the ChromeDriver executable
    chromedriver_path = '/path/to/chromedriver'
    username = 'gamodemy1@gmail.com'
    password = getpass.getpass('Enter your password: ')

    sender_email = 'yashpratapr@gmail.com'
    sender_password = 'ayuobdfxotqbjwog'
    recipient_email = 'jainendra116@gmail.com'
    SMTP_SERVER = 'smtp.gmail.com'
    SMTP_PORT = 587

    linkedin_scraper = LinkedInScraper(chromedriver_path)

    while True:
        # Scrape the current data
        current_data = linkedin_scraper.scrape_unread_data(username, password)
        if current_data is not None:
            workbook_filename = 'linkedin_data.xlsx'
            # Retrieve the previous data from the Excel file
            previous_data = retrieve_previous_data(workbook_filename)
            #comparison
            comparison = findComparison(current_data, previous_data)
            # Update the Excel file with the current data
            update_excel_data(workbook_filename, [
                username,
                time.strftime("%d-%B, %H:%M"),
                current_data[0],
                current_data[1],
                comparison["Unread Messages:"],
                comparison["Unread Notifications:"]
            ])
            # Generate the email body
            email_body = generate_email_body(current_data, previous_data)
            subject = 'LinkedIn Unread Messages and Notifications'
            # Send the email
            send_email(sender_email, sender_password, recipient_email, SMTP_SERVER, SMTP_PORT, subject, email_body)
        time.sleep(3 * 60 * 60)  # Sleep for 3 hours
